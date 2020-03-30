from gevent import monkey
monkey.patch_all()
import gevent
from gevent.pool import Pool
import requests
from bs4 import BeautifulSoup
import lxml
import re
import openpyxl
from 获取代理IP.get_proxies import get_proxiex_list
import random


def get_data(city,year,month,html_set):
    proxies=random.choice(proxies_list)
    url = 'http://www.tianqihoubao.com/lishi/{0}/month/{1}{2}.html'.format(city, year, month)
    req = requests.get(url=url,proxies=proxies,timeout=30)
    print(req.url)
    html_set.add(req.text)
    return html_set


def parse_data(html,data_list):
    soup = BeautifulSoup(html, 'lxml')
    city=soup.find('div',class_='wdetail').find('h1').text.split('历史')[0].strip()
    table = soup.find('table', class_='b')
    tr_list = table.find_all('tr')
    for tr in tr_list[1:]:
        # 获得每天的数据列表
        td_list = tr.find_all('td')
        rq = td_list[0].text.strip()  # 日期
        tqzk = re.sub(r'\s*', '', td_list[1].text.strip())  # 天气状况
        qw = re.sub(r'\s*', '', td_list[2].text.strip())  # 气温
        fx = re.sub(r'\s*', '', td_list[3].text.strip())  # 风力风向
        row=(city,rq,tqzk,qw,fx)
        data_list.append(row)
    return data_list


if __name__=='__main__':
    proxies_list=get_proxiex_list()
    pool=Pool(20)
    task_list=[]
    data_list = []
    html_set=set()
    city_list = ['shenyang','dalian','changchun','haerbin','beijing',
                 'huhehaote','shijiazhuang','taiyuan','tangshan','jinan',
                 'qingdao','zhengzhou','hefei','lanzhou','wulumuqi','xian','yinchuan',
                 'chengdu','guiyang','kunming','wuhan','changsha','nanchang','shanghai',
                 'hangzhou','nanjing','haikou','nanning','xiamen','shenzhen']
    month_list = ['01', '02', '03', '04', '05', '06', '07', '08','09','10','11','12']
    for city in city_list:
        for month in month_list:
            try:
                p=pool.spawn(get_data,city,'2018',month,html_set)
                task_list.append(p)
            except Exception as e:
                print(e)
    gevent.joinall(task_list)

    wb = openpyxl.Workbook()
    wb.save('城市历史天气预报.xlsx')
    wb=openpyxl.load_workbook('城市历史天气预报.xlsx')
    sht=wb.get_active_sheet()
    sht.cell(1,1,'城市')
    sht.cell(1, 2, '日期')
    sht.cell(1, 3, '天气状况')
    sht.cell(1, 4, '气温')
    sht.cell(1, 5, '风力风向')

    r=2
    try:
        for html in list(html_set):
            data_list=parse_data(html, data_list)
        for one in data_list:
            sht.cell(r, 1, one[0])
            sht.cell(r, 2, one[1])
            sht.cell(r, 3, one[2])
            sht.cell(r, 4, one[3])
            sht.cell(r, 5, one[4])
            r += 1
            print(one[0])
    except Exception as e:
        print(e)

    wb.save('城市历史天气预报.xlsx')