import requests
from bs4 import BeautifulSoup
import lxml
import re
import openpyxl


wb=openpyxl.Workbook()
wb.save('城市历史天气预报.xlsx')


while True:
    try:
        city=input('请输入城市拼音：')
        year=input('请输入年份：') #最小年份为2011年
        month_list=['01','02','03','04','05','06','07','08','09','10','11','12']
        #month_list=['01','02']
        title=False
        sheet_names=wb.get_sheet_names()
        sht=wb.get_sheet_by_name(sheet_names[0])
        sht.cell(1,1,'日期')
        sht.cell(1,2,'天气状况')
        sht.cell(1,3,'气温')
        sht.cell(1,4,'风力风向')

        r=2
        for month in month_list:
            url='http://www.tianqihoubao.com/lishi/{0}/month/{1}{2}.html'.format(city,year,month)
            req=requests.get(url)
            html=req.text
            soup=BeautifulSoup(html,'lxml')
            if not title :
                title_sheet=soup.find('title').text.split('年')[0].strip()  #工作表名称


            table=soup.find('table',class_='b')
            tr_list=table.find_all('tr')
            for tr in tr_list[1:]:
                #获得每天的数据列表
                td_list=tr.find_all('td')
                rq=td_list[0].text.strip()    #日期
                tqzk=re.sub(r'\s*','',td_list[1].text.strip())   #天气状况
                qw=re.sub(r'\s*','',td_list[2].text.strip())    #气温
                fx=re.sub(r'\s*','',td_list[3].text.strip())    #风力风向


                sht.cell(r,1,rq)
                sht.cell(r,2,tqzk)
                sht.cell(r,3,qw)
                sht.cell(r,4,fx)
                r+=1

        sht.title=title_sheet
    except Exception as e:
        print(e)

    isgo=input('是否继续（n/y）:')
    if isgo=='n':
        break
        
    sht=wb.create_sheet(index=0)
    
                    


wb.save('城市历史天气预报.xlsx')
            
